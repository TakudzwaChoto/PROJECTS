import matplotlib.pyplot as plt
import numpy as np
from openpyxl.reader.excel import load_workbook
from pandas.io.formats.printing import pprint_thing

from new import df

'''
x = np.array([1, 2, 3, 4, 5, 6, 7, 8])
y = np.array([1, 4, 9, 16, 7, 11, 23, 18])
sizes = np.array([20,50,100,200,500,1000,60,90])
plt.scatter(x, y, s=sizes)
plt.show()

#Blue Camp
soldiers = []
class soldiers:
    def Blue_Camp(self=0):
       x = np.array([5,7,8,7,2,17,2,9,4,11,12,9,6])
       y = np.array([99,86,87,88,111,86,103,87,94,78,77,85,86])
       plt.title('Blue Camp')
       plt.scatter(x, y, color = 'blue')


#Red Camp
    def Red_Camp(self=0):
       x = np.array([2,2,8,1,15,8,12,9,7,3,11,4,7,14,12])
       y = np.array([100,105,84,105,90,99,90,95,94,100,79,112,91,80,85])
       plt.title('Red Camp')


    for x in soldiers:
        plt.scatter(x, y, color='red')
        Blue_Camp.append(Red_Camp())


plt.show()

x = np.array([5,7,8,7,2,17,2,9,4,11,12,9,6])
y = np.array([99,86,87,88,111,86,103,87,94,78,77,85,86])
colors = np.array([0, 10, 20, 30, 40, 45, 50, 55, 60, 70, 80, 90, 100])

plt.scatter(x, y, c=colors, cmap='viridis')

plt.colorbar()

plt.show()



# initialize x and y coordinates
x = [0.1, 0.2, 0.3, 0.4, 0.5]
y = [6.2, -8.4, 8.5, 9.2, -6.3]

# set the title of a plot
plt.title("Connected Scatterplot points with lines")

# plot scatter plot with x and y data
plt.scatter(x, y)

# plot with x and y data
plt.plot(x, y)
plt.show()

import matplotlib.pyplot as plt
import  networkx as nx
# 导入模块函数
G = nx.cubical_graph() # 生成一个正则图（3-regular Platonic Cubical graph）
plt.subplot(121) # 绘制子图
nx.draw(G)
plt.subplot(122)
nx.draw(G,pos=nx.circular_layout(G),nodecolor='r',edge_color='b')

from matplotlib import pyplot as plt

import networkx as nx

G=nx.path_graph(5)

path=nx.single_source_shortest_path(G,2)

length=nx.single_source_shortest_path_length(G,2)

print(path)

print(length)

nx.draw_networkx(G)

plt.show()

import networkx as nx

nodes=[0,1,2,3,4]

edges=[(0,1,10),(0,3,30),(0,4,100),(1,2,50),(2,3,20),(2,4,10),(3,4,60)]

G=nx.Graph()

G.add_nodes_from(nodes)

G.add_weighted_edges_from(edges)

path=nx.single_source_dijkstra_path(G,4)

length=nx.single_source_dijkstra_path_length(G,4)

print(path)

print(length)

nx.draw_networkx(G)

plt.show()

import networkx as nx

net_grid = nx.Graph()  # 新建一个无向图 net_grid

# 网络图中的所有的节点，以列表的形式存储
list_net_nodes = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]

# 网络图中的所有的边，以列表的形式存储，由于是无向图，可以自动忽略掉重读的边
list_net_edges = [(1, 3), (3, 5), (5, 4), (4, 2), (2, 6),
                  (5, 7), (5, 8), (8, 6),
                  (7, 9), (8, 9), (6, 10),
                  (9, 11), (10, 12), (10, 13),
                  (11, 14), (12, 14), (12, 15),
                  (14, 16), (15, 16), (15, 17),
                  (16, 18), (17, 19),
                  (18, 20), (19, 20)]

# 在网络图中添加节点，直接从节点列表中读取所有的节点
net_grid.add_nodes_from(list_net_nodes)

# 在网络中添加边，直接从边列表中读取所有的边
net_grid.add_edges_from(list_net_edges)

# 画网络图
nx.draw(net_grid, with_labels=True)

# 显示
plt.show()

import networkx as nx

# tag names specifying what game info should be
# stored in the dict on each digraph edge
game_details=[
              "Event",
              "Date",
              "Result",
              "ECO",
              "Site"
             ]

def chess_pgn_graph(pgn_file="chess_masters_WCC.pgn.bz2"):
    """Read chess games in pgn format in pgn_file.

    Filenames ending in .gz or .bz2 will be uncompressed.

    Return the MultiDiGraph of players connected by a chess game.
    Edges contain game data in a dict.

    """
    try:# use networkx.utils._get_fh to uncompress
        # pgn file if required
        datafile=nx.utils._get_fh(pgn_file,mode='rb')
    except IOError:
        print ("Could not read file %s."%(pgn_file))
        raise
    G=nx.MultiDiGraph(weighted=False)
    game_info={}
    for line in datafile.read().splitlines():
        # check for tag pairs
        if len(line)>0 and line[0]=='[':
           line=line[1:-1] # remove extra quotes
           tag = line.split()[0]
           value=line[len(tag)+2:-1]
           if tag=='White':
              white=value.split(',')[0]
           elif tag=='Black':
              black=value.split(',')[0]
           elif tag in game_details:
              game_info[tag]=value
        # empty line after tag set indicates
        # we finished reading game info
        elif len(line)==0:
             if len(game_info)>0:
                 G.add_edge(white, black, **game_info)
                 game_info={}
    return G


if __name__ == '__main__':
    import networkx as nx
    import matplotlib.pyplot as plt

    plt.rcParams['text.usetex'] = False

    G=chess_pgn_graph()
    ngames=G.number_of_edges()
    nplayers=G.number_of_nodes()

    print ("Loaded %d chess games between %d players\n"
                   % (ngames,nplayers))

    # identify connected components
    # of the undirected version
    Gcc=nx.connected_component_subgraphs(G.to_undirected())
    if len(Gcc)>1:
        print ("Note the disconnected component consisting of:")
        print (Gcc[1].nodes())

    # find all games with B97 opening (as described in ECO)
    openings=set([game_info['ECO']
                  for (white,black,game_info) in G.edges(data=True)])
    print("\nFrom a total of %d different openings,"%len(openings))
    print ('the following games used the Sicilian opening')
    print ('with the Najdorff 7...Qb6 "Poisoned Pawn" variation.\n')

    for (white,black,game_info) in G.edges(data=True):
        if game_info['ECO']=='B97':
           print (white,"vs",black)
           for k,v in game_info.items():
               print( "   ",k,": ",v)
           print ("\n")


    plt.figure(figsize=(8,8))
    # make new undirected graph H without multi-edges
    H=nx.Graph(G)
    # edge width is proportional number of games played
    edgewidth=[]
    for (u,v,d) in H.edges(data=True):
        edgewidth.append(len(G.get_edge_data(u,v)))

    # node size is proportional to number of games won
    wins=dict.fromkeys(G.nodes(),0.0)
    for (u,v,d) in G.edges(data=True):
        r=d['Result'].split('-')
        if r[0]=='1':
            wins[u]+=1.0
        elif r[0]=='1/2':
            wins[u]+=0.5
            wins[v]+=0.5
        else:
            wins[v]+=1.0

    print (H.edges(data=True))
    A=nx.to_numpy_matrix(H)
    try:
        pos=nx.graphviz_layout(H)
    except:
        pos=nx.spring_layout(H,iterations=20)

    nx.draw_networkx_edges(H,pos,alpha=0.3,width=edgewidth, edge_color='m')
    nodesize=[wins[v]*50 for v in H]
    nx.draw_networkx_nodes(H,pos,node_size=nodesize,node_color='w',alpha=0.4)
    nx.draw_networkx_edges(H,pos,alpha=0.4,node_size=0,width=1,edge_color='k')
    nx.draw_networkx_labels(H,pos,fontsize=14)
    font = {'fontname'   : 'Helvetica',
            'color'      : 'k',
            'fontweight' : 'bold',
            'fontsize'   : 14}
    plt.title("World Chess Championship Games: 1886 - 1985", font)

    # change font and write text (using data coordinates)
    font = {'fontname'   : 'Helvetica',
    'color'      : 'r',
    'fontweight' : 'bold',
    'fontsize'   : 14}

    plt.text(0.5, 0.97, "edge width = # games played",
             horizontalalignment='center',
             transform=plt.gca().transAxes)
    plt.text(0.5, 0.94,  "node size = # games won",
             horizontalalignment='center',
             transform=plt.gca().transAxes)

    plt.axis('off')
    plt.savefig("chess_masters.png",dpi=75)
    print ("Wrote chess_masters.png")
    plt.show() # display
    

import matplotlib.pyplot as plt
import networkx as nx

# tag names specifying what game info should be
# stored in the dict on each digraph edge
game_details = ["Event", "Date", "Result", "ECO", "Site"]


def chess_pgn_graph(pgn_file="chess_masters_WCC.pgn.bz2"):
    """Read chess games in pgn format in pgn_file.

    Filenames ending in .bz2 will be uncompressed.

    Return the MultiDiGraph of players connected by a chess game.
    Edges contain game data in a dict.

    """
    import bz2

    G = nx.MultiDiGraph()
    game = {}
    with bz2.BZ2File(pgn_file) as datafile:
        lines = [line.decode().rstrip("\r\n") for line in datafile]
    for line in lines:
        if line.startswith("["):
            tag, value = line[1:-1].split(" ", 1)
            game[str(tag)] = value.strip('"')
        else:
            # empty line after tag set indicates
            # we finished reading game info
            if game:
                white = game.pop("White")
                black = game.pop("Black")
                G.add_edge(white, black, **game)
                game = {}
    return G


G = chess_pgn_graph()

print(
    f"Loaded {G.number_of_edges()} chess games between {G.number_of_nodes()} players\n"
)

# identify connected components of the undirected version
H = G.to_undirected()
Gcc = [H.subgraph(c) for c in nx.connected_components(H)]
if len(Gcc) > 1:
    print(f"Note the disconnected component consisting of:\n{Gcc[1].nodes()}")

# find all games with B97 opening (as described in ECO)
openings = {game_info["ECO"] for (white, black, game_info) in G.edges(data=True)}
print(f"\nFrom a total of {len(openings)} different openings,")
print("the following games used the Sicilian opening")
print('with the Najdorff 7...Qb6 "Poisoned Pawn" variation.\n')

for (white, black, game_info) in G.edges(data=True):
    if game_info["ECO"] == "B97":
        summary = f"{white} vs {black}\n"
        for k, v in game_info.items():
            summary += f"   {k}: {v}\n"
        summary += "\n"
        print(summary)

# make new undirected graph H without multi-edges
H = nx.Graph(G)

# edge width is proportional number of games played
edgewidth = [len(G.get_edge_data(u, v)) for u, v in H.edges()]

# node size is proportional to number of games won
wins = dict.fromkeys(G.nodes(), 0.0)
for (u, v, d) in G.edges(data=True):
    r = d["Result"].split("-")
    if r[0] == "1":
        wins[u] += 1.0
    elif r[0] == "1/2":
        wins[u] += 0.5
        wins[v] += 0.5
    else:
        wins[v] += 1.0
nodesize = [wins[v] * 50 for v in H]

# Generate layout for visualization
pos = nx.kamada_kawai_layout(H)
# Manual tweaking to limit node label overlap in the visualization
pos["Reshevsky, Samuel H"] += (0.05, -0.10)
pos["Botvinnik, Mikhail M"] += (0.03, -0.06)
pos["Smyslov, Vassily V"] += (0.05, -0.03)

fig, ax = plt.subplots(figsize=(12, 12))
# Visualize graph components
nx.draw_networkx_edges(H, pos, alpha=0.3, width=edgewidth, edge_color="m")
nx.draw_networkx_nodes(H, pos, node_size=nodesize, node_color="#210070", alpha=0.9)
label_options = {"ec": "k", "fc": "white", "alpha": 0.7}
nx.draw_networkx_labels(H, pos, font_size=14, bbox=label_options)

# Title/legend
font = {"fontname": "Helvetica", "color": "k", "fontweight": "bold", "fontsize": 14}
ax.set_title("World Chess Championship Games: 1886 - 1985", font)
# Change font color for legend
font["color"] = "r"

ax.text(
    0.80,
    0.10,
    "edge width = # games played",
    horizontalalignment="center",
    transform=ax.transAxes,
    fontdict=font,
)
ax.text(
    0.80,
    0.06,
    "node size = # games won",
    horizontalalignment="center",
    transform=ax.transAxes,
    fontdict=font,
)

# Resize figure for label readibility
ax.margins(0.1, 0.05)
fig.tight_layout()
plt.axis("off")
plt.show()

#############################################
import matplotlib.pyplot as plt
import networkx as nx

# A rainbow color mapping using matplotlib's tableau colors
node_dist_to_color = {
    1: "tab:red",
    2: "tab:orange",
    3: "tab:olive",
    4: "tab:green",
    5: "tab:blue",
    6: "tab:purple",
}

# Create a complete graph with an odd number of nodes
nnodes = 13
G = nx.complete_graph(nnodes)

# A graph with (2n + 1) nodes requires n colors for the edges
n = (nnodes - 1) // 2
ndist_iter = list(range(1, n + 1))

# Take advantage of circular symmetry in determining node distances
ndist_iter += ndist_iter[::-1]


def cycle(nlist, n):
    return nlist[-n:] + nlist[:-n]


# Rotate nodes around the circle and assign colors for each edge based on
# node distance
nodes = list(G.nodes())
for i, nd in enumerate(ndist_iter):
    for u, v in zip(nodes, cycle(nodes, i + 1)):
        G[u][v]["color"] = node_dist_to_color[nd]

pos = nx.circular_layout(G)
# Create a figure with 1:1 aspect ratio to preserve the circle.
fig, ax = plt.subplots(figsize=(8, 8))
node_opts = {"node_size": 500, "node_color": "w", "edgecolors": "k", "linewidths": 2.0}
nx.draw_networkx_nodes(G, pos, **node_opts)
nx.draw_networkx_labels(G, pos, font_size=14)
# Extract color from edge data
edge_colors = [edgedata["color"] for _, _, edgedata in G.edges(data=True)]
nx.draw_networkx_edges(G, pos, width=2.0, edge_color=edge_colors)

ax.set_axis_off()
fig.tight_layout()
plt.show()


import matplotlib.pyplot as plt
import networkx as nx

# A rainbow color mapping using matplotlib's tableau colors
node_dist_to_color = {
    1: "tab:red",
    2: "tab:orange",
    3: "tab:olive",
    4: "tab:green",
    5: "tab:blue",
    6: "tab:purple",
}

# Create a complete graph with an odd number of nodes
nnodes = 13
G = nx.complete_graph(nnodes)

# A graph with (2n + 1) nodes requires n colors for the edges
n = (nnodes - 1) // 2
ndist_iter = list(range(1, n + 1))

# Take advantage of circular symmetry in determining node distances
ndist_iter += ndist_iter[::-1]


def cycle(nlist, n):
    return nlist[-n:] + nlist[:-n]


# Rotate nodes around the circle and assign colors for each edge based on
# node distance
nodes = list(G.nodes())
for i, nd in enumerate(ndist_iter):
    for u, v in zip(nodes, cycle(nodes, i + 1)):
        G[u][v]["color"] = node_dist_to_color[nd]

pos = nx.circular_layout(G)
# Create a figure with 1:1 aspect ratio to preserve the circle.
fig, ax = plt.subplots(figsize=(8, 8))
node_opts = {"node_size": 500, "node_color": "w", "edgecolors": "k", "linewidths": 2.0}
nx.draw_networkx_nodes(G, pos, **node_opts)
nx.draw_networkx_labels(G, pos, font_size=14)
# Extract color from edge data
edge_colors = [edgedata["color"] for _, _, edgedata in G.edges(data=True)]
nx.draw_networkx_edges(G, pos, width=2.0, edge_color=edge_colors)

ax.set_axis_off()
fig.tight_layout()
plt.show()

from random import sample
import networkx as nx
import matplotlib.pyplot as plt

# Gold standard data of positive gene functional associations
# from https://www.inetbio.org/wormnet/downloadnetwork.php
G = nx.read_edgelist("WormNet.v3.benchmark.txt")

# remove randomly selected nodes (to make example fast)
num_to_remove = int(len(G) / 1.5)
nodes = sample(list(G.nodes), num_to_remove)
G.remove_nodes_from(nodes)

# remove low-degree nodes
low_degree = [n for n, d in G.degree() if d < 10]
G.remove_nodes_from(low_degree)

# largest connected component
components = nx.connected_components(G)
largest_component = max(components, key=len)
H = G.subgraph(largest_component)

# compute centrality
centrality = nx.betweenness_centrality(H, k=10, endpoints=True)

# compute community structure
lpc = nx.community.label_propagation_communities(H)
community_index = {n: i for i, com in enumerate(lpc) for n in com}

#### draw graph ####
fig, ax = plt.subplots(figsize=(20, 15))
pos = nx.spring_layout(H, k=0.15, seed=4572321)
node_color = [community_index[n] for n in H]
node_size = [v * 20000 for v in centrality.values()]
nx.draw_networkx(
    H,
    pos=pos,
    with_labels=False,
    node_color=node_color,
    node_size=node_size,
    edge_color="gainsboro",
    alpha=0.4,
)

# Title/legend
font = {"color": "k", "fontweight": "bold", "fontsize": 20}
ax.set_title("Gene functional association network (C. elegans)", font)
# Change font color for legend
font["color"] = "r"

ax.text(
    0.80,
    0.10,
    "node color = community structure",
    horizontalalignment="center",
    transform=ax.transAxes,
    fontdict=font,
)
ax.text(
    0.80,
    0.06,
    "node size = betweeness centrality",
    horizontalalignment="center",
    transform=ax.transAxes,
    fontdict=font,
)

# Resize figure for label readibility
ax.margins(0.1, 0.05)
fig.tight_layout()
plt.axis("off")
plt.show()

import networkx as nx
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

# The graph to visualize
G = nx.cycle_graph(20)

# 3d spring layout
pos = nx.spring_layout(G, dim=3, seed=779)
# Extract node and edge positions from the layout
node_xyz = np.array([pos[v] for v in sorted(G)])
edge_xyz = np.array([(pos[u], pos[v]) for u, v in G.edges()])

# Create the 3D figure
fig = plt.figure()
ax = fig.add_subplot(111, projection="3d")

# Plot the nodes - alpha is scaled by "depth" automatically
ax.scatter(*node_xyz.T, s=100, ec="w")

# Plot the edges
for vizedge in edge_xyz:
    ax.plot(*vizedge.T, color="tab:gray")


def _format_axes(ax):
    """Visualization options for the 3D axes."""
    # Turn gridlines off
    ax.grid(False)
    # Suppress tick labels
    for dim in (ax.xaxis, ax.yaxis, ax.zaxis):
        dim.set_ticks([])
    # Set axes labels
    ax.set_xlabel("x")
    ax.set_ylabel("y")
    ax.set_zlabel("z")


_format_axes(ax)
fig.tight_layout()
plt.show()

import matplotlib.pyplot as plt
import networkx as nx

# tag names specifying what game info should be
# stored in the dict on each digraph edge
game_details = ["Event", "Date", "Result", "ECO", "Site"]


def chess_pgn_graph(pgn_file="chess_masters_WCC.pgn.bz2"):
    """Read chess games in pgn format in pgn_file.

    Filenames ending in .bz2 will be uncompressed.

    Return the MultiDiGraph of players connected by a chess game.
    Edges contain game data in a dict.

    """
    import bz2

    G = nx.MultiDiGraph()
    game = {}
    with bz2.BZ2File(pgn_file) as datafile:
        lines = [line.decode().rstrip("\r\n") for line in datafile]
    for line in lines:
        if line.startswith("["):
            tag, value = line[1:-1].split(" ", 1)
            game[str(tag)] = value.strip('"')
        else:
            # empty line after tag set indicates
            # we finished reading game info
            if game:
                white = game.pop("White")
                black = game.pop("Black")
                G.add_edge(white, black, **game)
                game = {}
    return G


G = chess_pgn_graph()

print(
    f"Loaded {G.number_of_edges()} chess games between {G.number_of_nodes()} players\n"
)

# identify connected components of the undirected version
H = G.to_undirected()
Gcc = [H.subgraph(c) for c in nx.connected_components(H)]
if len(Gcc) > 1:
    print(f"Note the disconnected component consisting of:\n{Gcc[1].nodes()}")

# find all games with B97 opening (as described in ECO)
openings = {game_info["ECO"] for (white, black, game_info) in G.edges(data=True)}
print(f"\nFrom a total of {len(openings)} different openings,")
print("the following games used the Sicilian opening")
print('with the Najdorff 7...Qb6 "Poisoned Pawn" variation.\n')

for (white, black, game_info) in G.edges(data=True):
    if game_info["ECO"] == "B97":
        summary = f"{white} vs {black}\n"
        for k, v in game_info.items():
            summary += f"   {k}: {v}\n"
        summary += "\n"
        print(summary)

# make new undirected graph H without multi-edges
H = nx.Graph(G)

# edge width is proportional number of games played
edgewidth = [len(G.get_edge_data(u, v)) for u, v in H.edges()]

# node size is proportional to number of games won
wins = dict.fromkeys(G.nodes(), 0.0)
for (u, v, d) in G.edges(data=True):
    r = d["Result"].split("-")
    if r[0] == "1":
        wins[u] += 1.0
    elif r[0] == "1/2":
        wins[u] += 0.5
        wins[v] += 0.5
    else:
        wins[v] += 1.0
nodesize = [wins[v] * 50 for v in H]

# Generate layout for visualization
pos = nx.kamada_kawai_layout(H)
# Manual tweaking to limit node label overlap in the visualization
pos["Reshevsky, Samuel H"] += (0.05, -0.10)
pos["Botvinnik, Mikhail M"] += (0.03, -0.06)
pos["Smyslov, Vassily V"] += (0.05, -0.03)

fig, ax = plt.subplots(figsize=(12, 12))
# Visualize graph components
nx.draw_networkx_edges(H, pos, alpha=0.3, width=edgewidth, edge_color="m")
nx.draw_networkx_nodes(H, pos, node_size=nodesize, node_color="#210070", alpha=0.9)
label_options = {"ec": "k", "fc": "white", "alpha": 0.7}
nx.draw_networkx_labels(H, pos, font_size=14, bbox=label_options)

# Title/legend
font = {"fontname": "Helvetica", "color": "k", "fontweight": "bold", "fontsize": 14}
ax.set_title("World Chess Championship Games: 1886 - 1985", font)
# Change font color for legend
font["color"] = "r"

ax.text(
    0.80,
    0.10,
    "edge width = # games played",
    horizontalalignment="center",
    transform=ax.transAxes,
    fontdict=font,
)
ax.text(
    0.80,
    0.06,
    "node size = # games won",
    horizontalalignment="center",
    transform=ax.transAxes,
    fontdict=font,
)

# Resize figure for label readibility
ax.margins(0.1, 0.05)
fig.tight_layout()
plt.axis("off")
plt.show()

import matplotlib.pyplot as plt
import networkx as nx

# tag names specifying what game info should be
# stored in the dict on each digraph edge
game_details = ["Event", "Date", "Result", "ECO", "Site"]


def chess_pgn_graph(pgn_file="chess_masters_WCC.pgn.bz2"):
    """Read chess games in pgn format in pgn_file.

    Filenames ending in .bz2 will be uncompressed.

    Return the MultiDiGraph of players connected by a chess game.
    Edges contain game data in a dict.

    """
    import bz2

    G = nx.MultiDiGraph()
    game = {}
    with bz2.BZ2File(pgn_file) as datafile:
        lines = [line.decode().rstrip("\r\n") for line in datafile]
    for line in lines:
        if line.startswith("["):
            tag, value = line[1:-1].split(" ", 1)
            game[str(tag)] = value.strip('"')
        else:
            # empty line after tag set indicates
            # we finished reading game info
            if game:
                white = game.pop("White")
                black = game.pop("Black")
                G.add_edge(white, black, **game)
                game = {}
    return G


G = chess_pgn_graph()

print(
    f"Loaded {G.number_of_edges()} chess games between {G.number_of_nodes()} players\n"
)

# identify connected components of the undirected version
H = G.to_undirected()
Gcc = [H.subgraph(c) for c in nx.connected_components(H)]
if len(Gcc) > 1:
    print(f"Note the disconnected component consisting of:\n{Gcc[1].nodes()}")

# find all games with B97 opening (as described in ECO)
openings = {game_info["ECO"] for (white, black, game_info) in G.edges(data=True)}
print(f"\nFrom a total of {len(openings)} different openings,")
print("the following games used the Sicilian opening")
print('with the Najdorff 7...Qb6 "Poisoned Pawn" variation.\n')

for (white, black, game_info) in G.edges(data=True):
    if game_info["ECO"] == "B97":
        summary = f"{white} vs {black}\n"
        for k, v in game_info.items():
            summary += f"   {k}: {v}\n"
        summary += "\n"
        print(summary)

# make new undirected graph H without multi-edges
H = nx.Graph(G)

# edge width is proportional number of games played
edgewidth = [len(G.get_edge_data(u, v)) for u, v in H.edges()]

# node size is proportional to number of games won
wins = dict.fromkeys(G.nodes(), 0.0)
for (u, v, d) in G.edges(data=True):
    r = d["Result"].split("-")
    if r[0] == "1":
        wins[u] += 1.0
    elif r[0] == "1/2":
        wins[u] += 0.5
        wins[v] += 0.5
    else:
        wins[v] += 1.0
nodesize = [wins[v] * 50 for v in H]

# Generate layout for visualization
pos = nx.kamada_kawai_layout(H)
# Manual tweaking to limit node label overlap in the visualization
pos["Reshevsky, Samuel H"] += (0.05, -0.10)
pos["Botvinnik, Mikhail M"] += (0.03, -0.06)
pos["Smyslov, Vassily V"] += (0.05, -0.03)

fig, ax = plt.subplots(figsize=(12, 12))
# Visualize graph components
nx.draw_networkx_edges(H, pos, alpha=0.3, width=edgewidth, edge_color="m")
nx.draw_networkx_nodes(H, pos, node_size=nodesize, node_color="#210070", alpha=0.9)
label_options = {"ec": "k", "fc": "white", "alpha": 0.7}
nx.draw_networkx_labels(H, pos, font_size=14, bbox=label_options)

# Title/legend
font = {"fontname": "Helvetica", "color": "k", "fontweight": "bold", "fontsize": 14}
ax.set_title("World Chess Championship Games: 1886 - 1985", font)
# Change font color for legend
font["color"] = "r"

ax.text(
    0.80,
    0.10,
    "edge width = # games played",
    horizontalalignment="center",
    transform=ax.transAxes,
    fontdict=font,
)
ax.text(
    0.80,
    0.06,
    "node size = # games won",
    horizontalalignment="center",
    transform=ax.transAxes,
    fontdict=font,
)

# Resize figure for label readibility
ax.margins(0.1, 0.05)
fig.tight_layout()
plt.axis("off")
plt.show()

import matplotlib.pyplot as plt
import networkx as nx

G = nx.Graph()

G.add_edge("a", "b", weight=0.6)
G.add_edge("a", "c", weight=0.2)
G.add_edge("c", "d", weight=0.1)
G.add_edge("c", "e", weight=0.7)
G.add_edge("c", "f", weight=0.9)
G.add_edge("a", "d", weight=0.3)

elarge = [(u, v) for (u, v, d) in G.edges(data=True) if d["weight"] > 0.5]
esmall = [(u, v) for (u, v, d) in G.edges(data=True) if d["weight"] <= 0.5]

pos = nx.spring_layout(G, seed=7)  # positions for all nodes - seed for reproducibility

# nodes
nx.draw_networkx_nodes(G, pos, node_size=700)

# edges
nx.draw_networkx_edges(G, pos, edgelist=elarge, width=6)
nx.draw_networkx_edges(
    G, pos, edgelist=esmall, width=6, alpha=0.5, edge_color="b", style="dashed"
)

# node labels
nx.draw_networkx_labels(G, pos, font_size=20, font_family="sans-serif")
# edge weight labels
edge_labels = nx.get_edge_attributes(G, "weight")
nx.draw_networkx_edge_labels(G, pos, edge_labels)

ax = plt.gca()
ax.margins(0.08)
plt.axis("off")
plt.tight_layout()
plt.show()


import itertools
import matplotlib.pyplot as plt
import networkx as nx

subset_sizes = [5, 5, 4, 3, 2, 4, 4, 3]
subset_color = [
    "blue",
    "blue",
    "blue",
    "blue",
    "red",
    "red",
    "red",
    "red",
]


def multilayered_graph(*subset_sizes):
    extents = nx.utils.pairwise(itertools.accumulate((0,) + subset_sizes))
    layers = [range(start, end) for start, end in extents]
    G = nx.Graph()
    for (i, layer) in enumerate(layers):
        G.add_nodes_from(layer, layer=i)
    for layer1, layer2 in nx.utils.pairwise(layers):
        G.add_edges_from(itertools.product(layer1, layer2))
    return G


G = multilayered_graph(*subset_sizes)
color = [subset_color[data["layer"]] for v, data in G.nodes(data=True)]
pos = nx.multipartite_layout(G, subset_key="layer")
plt.figure(figsize=(8, 8))
nx.draw(G, pos, node_color=color, with_labels=False)
plt.axis("equal")
plt.show()

# 导入相关依赖
from matplotlib import pyplot as plt
import networkx as nx
import numpy as np

# 生成随机数据
G = nx.erdos_renyi_graph(50, 0.5)

# 指定画布大小
plt.figure(figsize=(18, 18))

# 生成新的图
G_new = nx.Graph()

# 依据图中边的数量，生成同样长度的随机权重值
weightList = {}
for i in range(len(G.edges()) + 1):
    weightList[i] = np.random.rand()

# 将生成的随机权重复制给G_new图
i = 0
for edge in G.edges():
    i += 1
    G_new.add_edges_from([(edge[0], edge[1], {'weight': weightList[i]})])
# 绘制G_new图
nx.draw_networkx(G_new)
plt.show()
'''
from openpyxl import Workbook
from openpyxl.workbook import Workbook
import pandas as pd
import networkx as nx
import seaborn as sns

workbook = Workbook()
spreadsheet = workbook.active
#spreadsheet["A1"] = "id"
#spreadsheet["B1"] = "from"
#spreadsheet ["C1"] = "to"
#spreadsheet ["D1"] = "length(km)"
#workbook.save(filename= r"C:\Users\ADMIN\AppData\Local\Temp\BNZ.6378dd56c2c68f6.xlsx")
#workbook = load_workbook(filename= r"C:\Users\ADMIN\AppData\Local\Temp\BNZ.6378dd56c2c68f6.xlsx")
#sheetnames = workbook.sheetnames
#print(sheetnames)
#print(spreadsheet)

dataframe_excel = pd.read_excel(r'C:\Users\ADMIN\AppData\Local\Temp\BNZ.6378dcd9c2a8310\Annex1：node＆link.xlsx')
#dataframe_excel1 = pd.read_excel(r'C:\Users\ADMIN\AppData\Local\Temp\BNZ.6378dcd9c2a8310\Annex1：node＆link.xlsx')
print(dataframe_excel)
print(df.describe())
print(df.info())
print(df.max())
print(df.std())
print(df.head(10))
print(df.shape)
print(df.sample(10))
#print(df.groupby("camp").max()["A1"])
#dataframe_csv.columns = ['', 'b','id', 'from', 'to', 'length(km)']
#dataframe_csv = pd.read_excel(r'C:\Users\ADMIN\AppData\Local\Temp\BNZ.6378dd56c2c68f6.xlsx')
'''
#print(dataframe_csv)
sheet = workbook['Sheet1']
multiple_cells = sheet['', 'b','id', 'from', 'to', 'length(km)']
for row in multiple_cells:
    for cell in row:
        print(cell.value)

sheet = workbook['Sheet1']
for column in sheet.columns:
    print(column)

multiple_cells = sheet['A1':'B3']
for row in multiple_cells:
    for cell in row:
        print(cell.value)

for row in sheet.rows:
    print(row)

dt = dataframe_excel,dataframe_excel1

for g in dt :
    df.plot.scatter(dataframe_excel,dataframe_excel1)
    plt.tight_layout()
    plt.show()
    '''
for col in dataframe_excel.columns:
    print(col)

#for row in dataframe_excel.rows:
 #   print(row)
vars = pd.read_excel(r'C:\Users\ADMIN\AppData\Local\Temp\BNZ.6378dcd9c2a8310\Annex1：node＆link.xlsx')
print(vars)

x = list(vars['x'])
y = list(vars['y'])

'''
G=nx.Graph()
H= x
for k in H:
  nx.path_graph(10)
  G.add_nodes_from(H)
  nx.draw(G, with_labels=True)
  plt.show()
'''
plt.figure(figsize=(10,10))
plt.style.use('seaborn')
plt.scatter(x,y,marker="*",s=100,edgecolors="red",c="red" )
plt.title("Attack difficulty Scatter Plot")
plt.show()

'''

def subgraph():
    G = nx.Graph()
    # 转化为图结构
    for node in x:
        G.add_node(node)

    for link in y:
        G.add_edge(link[0], link[1])

        plt.subplot(211)
        nx.draw_networkx(G, with_labels=True)
        color =['y','g']
        subplot = [223,224]
        plt.show()
    # 打印连通子图
        for c in nx.connected_components(G):

         print("连通子图：", c)
       # 绘制子图
         subgraph = G.subgraph(c)
         plt.subplot(subplot[0])  # 第二整行
         nx.draw_networkx(subgraph, with_labels=True,node_color=color[0])
         color.pop(0)
         subplot.pop(0)
plt.show()
'''

